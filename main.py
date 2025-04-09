from fastapi import FastAPI, Depends, HTTPException, status, Query, Response, Request
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_, desc, func
from typing import List, Optional, Dict, Any
from datetime import timedelta, datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from fastapi.openapi.utils import get_openapi
from fastapi.routing import APIRouter
from fastapi.staticfiles import StaticFiles
from fastapi.openapi.docs import get_swagger_ui_html, get_redoc_html
import logging
from pydantic import BaseModel

import models
import schemas
import auth
from database import engine, get_db, filter_deleted
from auth import (
    get_current_active_user, 
    get_current_superadmin, 
    authenticate_user, 
    create_access_token,
    ACCESS_TOKEN_EXPIRE_MINUTES,
    get_password_hash,
    check_region_access,
    conditional_auth
)
from data_initializer import initialize_database
from logging_config import RequestLoggingMiddleware

# Configure logging to reduce noise from invalid HTTP requests
logging.getLogger("uvicorn.error").setLevel(logging.ERROR)
logging.getLogger("uvicorn.access").setLevel(logging.INFO)

# Create database tables
models.Base.metadata.create_all(bind=engine)

# Replace the FastAPI app initialization with this:
app = FastAPI(
    title="Project Management API",
    description="API for managing regional projects with user authentication",
    version="1.0.0",
    # Enable default docs endpoints
    docs_url="/docs",
    redoc_url="/redoc"
)

# Add this function to the main.py file, right after the imports
@app.middleware("http")
async def debug_request(request: Request, call_next):
    # Log request details for debugging
    print(f"Request path: {request.url.path}")
    print(f"Request method: {request.method}")
    print(f"Request headers: {request.headers}")
    
    # Continue processing the request
    response = await call_next(request)
    return response

# Modified trailing slash middleware to be more flexible
@app.middleware("http")
async def handle_trailing_slash(request: Request, call_next):
    """Middleware to handle trailing slashes in URLs."""
    path = request.url.path
    
    # Skip for static files, docs endpoints, and OpenAPI endpoints
    if path.startswith('/static/') or path in ['/docs', '/redoc', '/openapi.json']:
        return await call_next(request)
    
    # For API endpoints, be more flexible with trailing slashes
    if path != "/" and path != "":  # Don't process root path
        original_path = path
        original_url = str(request.url)
        
        # Try both with and without trailing slash
        try_paths = []
        if path.endswith('/'):
            try_paths = [path, path.rstrip('/')]
        else:
            try_paths = [path, f"{path}/"]
        
        # Try the original path first
        try:
            response = await call_next(request)
            if response.status_code != 404 and response.status_code != 405:
                return response
        except Exception:
            pass
        
        # If we got a 404 or 405, try the alternative path
        for try_path in try_paths:
            if try_path != original_path:
                # Modify the request path
                request.scope["path"] = try_path
                request.scope["raw_path"] = try_path.encode()
                
                # Update the full URL
                if original_url.endswith('/') and not try_path.endswith('/'):
                    request.scope["raw_path"] = original_url[:-1].encode()
                elif not original_url.endswith('/') and try_path.endswith('/'):
                    request.scope["raw_path"] = (original_url + '/').encode()
                
                try:
                    response = await call_next(request)
                    if response.status_code != 404 and response.status_code != 405:
                        return response
                    # Restore original path if we still get an error
                    request.scope["path"] = original_path
                    request.scope["raw_path"] = original_path.encode()
                except Exception:
                    # Restore original path if we get an exception
                    request.scope["path"] = original_path
                    request.scope["raw_path"] = original_path.encode()
    
    # Continue with normal processing if all alternatives failed
    return await call_next(request)

# Add CORS middleware
app.add_middleware(
  CORSMiddleware,
  allow_origins=["http://localhost:5173", "http://localhost:5174", "https://admin-panel-qq-eco-social.netlify.app", "https://qq-ekonomika-social.netlify.app", "https://localhost:5173", "https://localhost:5174"],  # List specific origins instead of "*"
  allow_credentials=True,
  allow_methods=["*"],
  allow_headers=["*"],
  expose_headers=["Authorization", "Content-Disposition"],
)

# Add RequestLoggingMiddleware
app.add_middleware(RequestLoggingMiddleware)

# Create routers for different categories
auth_router = APIRouter(tags=["Authentication"])
users_router = APIRouter(tags=["User Management"])
regions_router = APIRouter(tags=["Region Management"])
projects_router = APIRouter(tags=["Project Management"])
authorities_router = APIRouter(tags=["Authority Management"])
statuses_router = APIRouter(tags=["Status Management"])

# Health check endpoint for load balancers and monitoring
@app.get("/health", include_in_schema=False)
async def health_check():
    return {"status": "healthy"}

# Add new endpoint to check API status
@app.get("/api-status", tags=["Status"])
async def api_status():
    return {
        "status": "online",
        "version": "1.0.0",
        "timestamp": datetime.utcnow().isoformat()
    }

# Handle OPTIONS requests explicitly to prevent invalid HTTP request warnings
@app.options("/{path:path}")
async def options_handler(request: Request, path: str):
    return Response(status_code=200)

# Define a model for JSON login
class LoginRequest(BaseModel):
    username: str
    password: str

# Add a new JSON-based login endpoint
@auth_router.post("/login", response_model=schemas.Token)
async def login_json(
    login_data: LoginRequest,
    db: Session = Depends(get_db)
):
    user = authenticate_user(db, login_data.username, login_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

# Authentication endpoints
@auth_router.post("/token", response_model=schemas.Token)
async def login_for_access_token(
    form_data: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db)
):
    user = authenticate_user(db, form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

# User management endpoints (superadmin only)
@users_router.post("/users", response_model=schemas.User)
@users_router.post("/users/", response_model=schemas.User)
async def create_user(
    user: schemas.UserCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    # Check if username already exists (including soft-deleted users)
    db_user = db.query(models.User).filter(
        models.User.username == user.username
    ).first()
    
    # Validate region assignment for non-superadmins
    if not user.is_superadmin and (not user.region_ids or len(user.region_ids) == 0):
        raise HTTPException(
            status_code=400, 
            detail="Regular admins must be assigned to at least one region"
        )
    
    if db_user and db_user.deleted_at is None:
        raise HTTPException(status_code=400, detail="Username already registered")
    elif db_user:
        # If user was soft-deleted, restore it with new data
        db_user.password_hash = get_password_hash(user.password)
        db_user.is_superadmin = user.is_superadmin
        db_user.is_active = user.is_active
        db_user.deleted_at = None
    else:
        # Create new user
        hashed_password = get_password_hash(user.password)
        db_user = models.User(
            username=user.username,
            password_hash=hashed_password,
            is_superadmin=user.is_superadmin,
            is_active=user.is_active
        )
        db.add(db_user)
        db.commit()  # Commit to get the user ID
    
    # Assign regions for non-superadmin users
    if not user.is_superadmin and user.region_ids:
        # Verify all region IDs exist and are not deleted
        regions = []
        for region_id in user.region_ids:
            region = db.query(models.Region).filter(
                models.Region.id == region_id,
                models.Region.deleted_at == None
            ).first()
            if region is None:
                raise HTTPException(status_code=404, detail=f"Region with ID {region_id} not found")
            regions.append(region)
        
        # Update user's regions
        db_user.regions = regions
    
    db.commit()
    db.refresh(db_user)
    return db_user

# User profile endpoint - always requires authentication
@users_router.get("/users/me", response_model=schemas.UserWithRegions)
@users_router.get("/users/me/", response_model=schemas.UserWithRegions)
async def read_users_me(current_user: models.User = Depends(get_current_active_user)):
    if current_user is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required for user data"
        )
    return current_user

@users_router.get("/users/me/regions", response_model=List[schemas.Region])
@users_router.get("/users/me/regions/", response_model=List[schemas.Region])
async def read_users_regions(current_user: models.User = Depends(get_current_active_user)):
    if current_user is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required for user data"
        )
    # Filter out deleted regions
    return [region for region in current_user.regions if region.deleted_at is None]

@users_router.get("/users", response_model=List[schemas.UserWithRegions])
@users_router.get("/users/", response_model=List[schemas.UserWithRegions])
async def read_users(
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    # This endpoint requires authentication since it's user-related
    if current_user is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required for user data"
        )
    
    # Only superadmins can see all users
    if not current_user.is_superadmin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    query = db.query(models.User)
    if not include_deleted:
        query = filter_deleted(query, models.User)
    users = query.all()
    return users

@users_router.get("/users/{user_id}", response_model=schemas.UserWithRegions)
async def read_user(
    user_id: int,
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    # This endpoint requires authentication since it's user-related
    if current_user is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required for user data"
        )
    
    # Only superadmins can see other users
    if not current_user.is_superadmin and current_user.id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    query = db.query(models.User).filter(models.User.id == user_id)
    if not include_deleted:
        query = filter_deleted(query, models.User)
    db_user = query.first()
    
    if db_user is None:
        raise HTTPException(status_code=404, detail="User not found")
    return db_user

# Support both with and without trailing slash for PUT
@users_router.put("/users/{user_id}", response_model=schemas.User)
@users_router.put("/users/{user_id}/", response_model=schemas.User)
async def update_user(
    user_id: int,
    user_update: schemas.UserUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    db_user = db.query(models.User).filter(
        models.User.id == user_id,
        models.User.deleted_at == None
    ).first()
    
    if db_user is None:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update user fields if provided
    if user_update.username is not None:
        # Check if username already exists
        existing_user = db.query(models.User).filter(
            models.User.username == user_update.username,
            models.User.id != user_id,
            models.User.deleted_at == None
        ).first()
        if existing_user:
            raise HTTPException(status_code=400, detail="Username already taken")
        db_user.username = user_update.username
    
    if user_update.password is not None:
        db_user.password_hash = get_password_hash(user_update.password)
    
    if user_update.is_active is not None:
        db_user.is_active = user_update.is_active
    
    db.commit()
    db.refresh(db_user)
    return db_user

# Add PATCH method for partial updates
@users_router.patch("/users/{user_id}", response_model=schemas.User)
@users_router.patch("/users/{user_id}/", response_model=schemas.User)
async def patch_user(
    user_id: int,
    user_update: schemas.UserUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    # Reuse the same implementation as PUT
    return await update_user(user_id, user_update, db, current_user)

@users_router.delete("/users/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
@users_router.delete("/users/{user_id}/", status_code=status.HTTP_204_NO_CONTENT)
async def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    db_user = db.query(models.User).filter(
        models.User.id == user_id,
        models.User.deleted_at == None
    ).first()
    
    if db_user is None:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Prevent deleting yourself
    if db_user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    # Soft delete
    db_user.soft_delete(db)
    return None

# Add endpoint to restore a deleted user
@users_router.post("/users/{user_id}/restore", response_model=schemas.User)
async def restore_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    db_user = db.query(models.User).filter(
        models.User.id == user_id,
        models.User.deleted_at != None
    ).first()
    
    if db_user is None:
        raise HTTPException(status_code=404, detail="Deleted user not found")
    
    # Restore user
    db_user.restore(db)
    return db_user

# User-Region management (superadmin only)
@users_router.put("/users/{user_id}/regions", response_model=schemas.UserWithRegions)
@users_router.put("/users/{user_id}/regions/", response_model=schemas.UserWithRegions)
async def update_user_regions(
    user_id: int,
    region_update: schemas.UserRegionUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    db_user = db.query(models.User).filter(
        models.User.id == user_id,
        models.User.deleted_at == None
    ).first()
    
    if db_user is None:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify all region IDs exist and are not deleted
    for region_id in region_update.region_ids:
        region = db.query(models.Region).filter(
            models.Region.id == region_id,
            models.Region.deleted_at == None
        ).first()
        if region is None:
            raise HTTPException(status_code=404, detail=f"Region with ID {region_id} not found")
    
    # Get all regions by IDs
    regions = db.query(models.Region).filter(
        models.Region.id.in_(region_update.region_ids),
        models.Region.deleted_at == None
    ).all()
    
    # Update user's regions
    db_user.regions = regions
    
    db.commit()
    db.refresh(db_user)
    return db_user

# Region endpoints
@regions_router.post("/regions", response_model=schemas.Region)
@regions_router.post("/regions/", response_model=schemas.Region)
async def create_region(
    region: schemas.RegionCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    # Check if a soft-deleted region with the same name exists
    db_region = db.query(models.Region).filter(
        models.Region.name == region.name
    ).first()
    
    if db_region and db_region.deleted_at is None:
        raise HTTPException(status_code=400, detail="Region with this name already exists")
    elif db_region:
        # If region was soft-deleted, restore it with new data
        db_region.stat_code = region.stat_code
        db_region.deleted_at = None
    else:
        # Create new region
        db_region = models.Region(**region.dict())
        db.add(db_region)
    
    db.commit()
    db.refresh(db_region)
    return db_region

# Update the read_regions function to handle None current_user
@regions_router.get("/regions", response_model=List[schemas.Region])
@regions_router.get("/regions/", response_model=List[schemas.Region])
async def read_regions(
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(conditional_auth)
):
    # This GET endpoint doesn't require authentication
    query = db.query(models.Region)
    if not include_deleted:
        query = filter_deleted(query, models.Region)
    regions = query.all()
    return regions

# Update the read_region function to handle None current_user
@regions_router.get("/regions/{region_id}", response_model=schemas.Region)
async def read_region(
    region_id: int,
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(conditional_auth)
):
    # This GET endpoint doesn't require authentication
    query = db.query(models.Region).filter(models.Region.id == region_id)
    if not include_deleted:
        query = filter_deleted(query, models.Region)
    db_region = query.first()
    
    if db_region is None:
        raise HTTPException(status_code=404, detail="Region not found")
    return db_region

@regions_router.put("/regions/{region_id}", response_model=schemas.Region)
@regions_router.put("/regions/{region_id}/", response_model=schemas.Region)
async def update_region(
    region_id: int,
    region_update: schemas.RegionUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    # PUT requests always require authentication
    if current_user is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required"
        )
    
    # Check if user has access to this region
    if not check_region_access(current_user, region_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to update this region"
        )
    
    db_region = db.query(models.Region).filter(
        models.Region.id == region_id,
        models.Region.deleted_at == None
    ).first()
    
    if db_region is None:
        raise HTTPException(status_code=404, detail="Region not found")
    
    # Update region fields if provided
    if region_update.name is not None:
        # Check if name already exists
        existing_region = db.query(models.Region).filter(
            models.Region.name == region_update.name,
            models.Region.id != region_id,
            models.Region.deleted_at == None
        ).first()
        if existing_region:
            raise HTTPException(status_code=400, detail="Region name already exists")
        db_region.name = region_update.name
    
    if region_update.stat_code is not None:
        db_region.stat_code = region_update.stat_code
    
    db.commit()
    db.refresh(db_region)
    return db_region

# Add PATCH method for partial updates
@regions_router.patch("/regions/{region_id}", response_model=schemas.Region)
@regions_router.patch("/regions/{region_id}/", response_model=schemas.Region)
async def patch_region(
    region_id: int,
    region_update: schemas.RegionUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    # Reuse the same implementation as PUT
    return await update_region(region_id, region_update, db, current_user)

@regions_router.delete("/regions/{region_id}", status_code=status.HTTP_204_NO_CONTENT)
@regions_router.delete("/regions/{region_id}/", status_code=status.HTTP_204_NO_CONTENT)
async def delete_region(
    region_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    db_region = db.query(models.Region).filter(
        models.Region.id == region_id,
        models.Region.deleted_at == None
    ).first()
    
    if db_region is None:
        raise HTTPException(status_code=404, detail="Region not found")
    
    # Check if region has associated projects
    projects = db.query(models.Project).filter(
        models.Project.region_id == region_id,
        models.Project.deleted_at == None
    ).count()
    
    if projects > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete region with {projects} active projects"
        )
    
    # Soft delete
    db_region.soft_delete(db)
    return None

# Add endpoint to restore a deleted region
@regions_router.post("/regions/{region_id}/restore", response_model=schemas.Region)
async def restore_region(
    region_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    db_region = db.query(models.Region).filter(
        models.Region.id == region_id,
        models.Region.deleted_at != None
    ).first()
    
    if db_region is None:
        raise HTTPException(status_code=404, detail="Deleted region not found")
    
    # Restore region
    db_region.restore(db)
    return db_region

# Project management endpoints
@projects_router.post("/projects", response_model=schemas.Project)
@projects_router.post("/projects/", response_model=schemas.Project)
async def create_project(
  project: schemas.ProjectCreate,
  db: Session = Depends(get_db),
  current_user: models.User = Depends(get_current_active_user)
):
  # Check if user is authenticated
  if current_user is None:
      raise HTTPException(
          status_code=status.HTTP_401_UNAUTHORIZED,
          detail="Authentication required",
          headers={"WWW-Authenticate": "Bearer"},
      )
  
  # Check if user has access to the region
  if not check_region_access(current_user, project.region_id):
      raise HTTPException(
          status_code=status.HTTP_403_FORBIDDEN,
          detail="Not authorized to create projects in this region"
      )
  
  # Check if region exists
  region = db.query(models.Region).filter(
      models.Region.id == project.region_id,
      models.Region.deleted_at == None
  ).first()
  if not region:
      raise HTTPException(status_code=404, detail="Region not found")
  
  # Check if authority exists
  authority = db.query(models.Authority).filter(
      models.Authority.id == project.authority_id,
      models.Authority.deleted_at == None
  ).first()
  if not authority:
      raise HTTPException(status_code=404, detail="Authority not found")
  
  # Check if status exists
  status = db.query(models.Status).filter(
      models.Status.id == project.status_id,
      models.Status.deleted_at == None
  ).first()
  if not status:
      raise HTTPException(status_code=404, detail="Status not found")
  
  try:
      # Create project
      db_project = models.Project(**project.dict())
      db.add(db_project)
      db.commit()
      db.refresh(db_project)
      return db_project
  except Exception as e:
      db.rollback()
      raise HTTPException(
          status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
          detail=f"Error creating project: {str(e)}"
      )

# Update the read_projects function to handle None current_user
@projects_router.get("/projects", response_model=List[schemas.Project])
@projects_router.get("/projects/", response_model=List[schemas.Project])
async def read_projects(
    region_id: Optional[int] = None,
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(conditional_auth)
):
    # This GET endpoint doesn't require authentication
    query = db.query(models.Project)
    
    # Filter by region if specified
    if region_id is not None:
        # For authenticated users, check region access
        if current_user is not None and not current_user.is_superadmin:
            accessible_region_ids = [region.id for region in current_user.regions if region.deleted_at is None]
            if region_id not in accessible_region_ids:
                # For authenticated users without access, return 403
                # For unauthenticated users, just filter by the requested region
                if current_user is not None:
                    raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail="Not authorized to view projects in this region"
                    )
        query = query.filter(models.Project.region_id == region_id)
    else:
        # If no region specified and user is authenticated but not superadmin, 
        # only show projects from regions they have access to
        if current_user is not None and not current_user.is_superadmin:
            accessible_region_ids = [region.id for region in current_user.regions if region.deleted_at is None]
            if not accessible_region_ids:
                return []  # Return empty list if user has no accessible regions
            query = query.filter(models.Project.region_id.in_(accessible_region_ids))
    
    # Filter out deleted projects unless include_deleted is True
    if not include_deleted:
        query = filter_deleted(query, models.Project)
    
    projects = query.all()
    return projects

# New endpoint to get the latest update timestamp across all projects
@projects_router.get("/projects/last_update", response_model=dict)
async def get_last_update_timestamp(
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(conditional_auth)
):
    """
    Get the latest update timestamp across all projects.
    Returns the most recent updated_at value from any project.
    """
    # Base query for non-deleted projects
    query = db.query(func.max(models.Project.updated_at)).filter(
        models.Project.deleted_at == None
    )
    
    # Filter by user's accessible regions if not superadmin
    if current_user is not None and not current_user.is_superadmin:
        accessible_region_ids = [region.id for region in current_user.regions if region.deleted_at is None]
        if not accessible_region_ids:
            return {"last_update": None}  # Return None if user has no accessible regions
        query = query.filter(models.Project.region_id.in_(accessible_region_ids))
    
    # Get the latest update timestamp
    latest_update = query.scalar()
    
    # Return the timestamp or None if no projects exist
    return {
        "last_update": latest_update
    }

# New endpoint to get the most recently updated projects
@projects_router.get("/projects/last-updates", response_model=List[schemas.ProjectLastUpdate])
async def get_last_project_updates(
    limit: int = 10,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(conditional_auth)
):
    """
    Get the most recently updated projects across all regions.
    Returns a list of projects sorted by update time (newest first).
    """
    # Base query for non-deleted projects
    query = db.query(
        models.Project.id.label("project_id"),
        models.Project.name.label("project_name"),
        models.Region.name.label("region_name"),
        models.Project.updated_at
    ).join(
        models.Region, models.Project.region_id == models.Region.id
    ).filter(
        models.Project.deleted_at == None,
        models.Region.deleted_at == None
    )
    
    # Filter by user's accessible regions if not superadmin
    if current_user is not None and not current_user.is_superadmin:
        accessible_region_ids = [region.id for region in current_user.regions if region.deleted_at is None]
        if not accessible_region_ids:
            return []  # Return empty list if user has no accessible regions
        query = query.filter(models.Project.region_id.in_(accessible_region_ids))
    
    # Order by most recently updated and limit results
    results = query.order_by(desc(models.Project.updated_at)).limit(limit).all()
    
    # Convert to list of ProjectLastUpdate objects
    return [
        schemas.ProjectLastUpdate(
            project_id=result.project_id,
            project_name=result.project_name,
            region_name=result.region_name,
            updated_at=result.updated_at
        ) for result in results
    ]

# Find the filter_projects function and replace it with this improved version
@projects_router.get("/projects/filter", response_model=List[schemas.Project])
async def filter_projects(
    region_id: Optional[int] = None,
    budget_min: Optional[float] = None,
    budget_max: Optional[float] = None,
    status_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(conditional_auth)
):
    """
    Filter projects by region, budget range, and status.
    All parameters are optional - if not provided, that filter will be ignored.
    """
    # Start with base query for non-deleted projects
    query = db.query(models.Project).filter(models.Project.deleted_at == None)
    
    # Apply region filter if provided and not None
    if region_id is not None:
        # For authenticated users, check region access
        if current_user is not None and not current_user.is_superadmin:
            accessible_region_ids = [region.id for region in current_user.regions if region.deleted_at is None]
            if region_id not in accessible_region_ids:
                # For authenticated users without access, return 403
                if current_user is not None:
                    raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail="Not authorized to view projects in this region"
                    )
        query = query.filter(models.Project.region_id == region_id)
    else:
        # If no region specified and user is authenticated but not superadmin,
        # only show projects from regions they have access to
        if current_user is not None and not current_user.is_superadmin:
            accessible_region_ids = [region.id for region in current_user.regions if region.deleted_at is None]
            if not accessible_region_ids:
                return []  # Return empty result if user has no accessible regions
            query = query.filter(models.Project.region_id.in_(accessible_region_ids))
    
    # Apply budget filters if provided and not None
    if budget_min is not None:
        query = query.filter(models.Project.budget_million >= budget_min)
    
    if budget_max is not None:
        query = query.filter(models.Project.budget_million <= budget_max)
    
    # Apply status filter if provided and not None
    if status_id is not None:
        query = query.filter(models.Project.status_id == status_id)
    
    # Get all projects matching the filters
    try:
        projects = query.all()
        return projects
    except Exception as e:
        # Log the error and return a more helpful error message
        print(f"Error filtering projects: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Error filtering projects: {str(e)}"
        )

# New endpoint for exporting projects to Excel
@projects_router.get("/projects/export")
async def export_projects_to_excel(
    region_id: Optional[int] = None,
    budget_min: Optional[float] = None,
    budget_max: Optional[float] = None,
    status_id: Optional[int] = None,
    initiator: Optional[str] = None,
    name: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(conditional_auth)
):
    # This GET endpoint doesn't require authentication
    # Start with base query
    query = db.query(models.Project).filter(models.Project.deleted_at == None)
    
    # Apply region filter and check access for authenticated users
    if region_id is not None:
        # For authenticated users, check region access
        if current_user is not None and not current_user.is_superadmin:
            accessible_region_ids = [region.id for region in current_user.regions if region.deleted_at is None]
            if region_id not in accessible_region_ids:
                # For authenticated users, return 403
                # For unauthenticated users, just filter by the requested region
                if current_user is not None:
                    raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail="Not authorized to export projects in this region"
                    )
        query = query.filter(models.Project.region_id == region_id)
    else:
        # If no region specified and user is authenticated but not superadmin,
        # only show projects from regions they have access to
        if current_user is not None and not current_user.is_superadmin:
            accessible_region_ids = [region.id for region in current_user.regions if region.deleted_at is None]
            if not accessible_region_ids:
                # Create empty Excel file
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Projects"
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                headers = {
                    'Content-Disposition': 'attachment; filename="projects.xlsx"',
                    'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'Access-Control-Expose-Headers': 'Content-Disposition'
                }
                return Response(content=output.getvalue(), headers=headers)
            
            query = query.filter(models.Project.region_id.in_(accessible_region_ids))
    
    # Apply other filters
    if budget_min is not None:
        query = query.filter(models.Project.budget_million >= budget_min)
    if budget_max is not None:
        query = query.filter(models.Project.budget_million <= budget_max)
    if status_id is not None:
        query = query.filter(models.Project.status_id == status_id)
    if initiator is not None:
        query = query.filter(models.Project.initiator.ilike(f"%{initiator}%"))
    if name is not None:
        query = query.filter(models.Project.name.ilike(f"%{name}%"))
    
    # Get all projects with their related data
    projects = query.all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"
    
    # Define headers
    headers = [
        "ID", "Region", "Initiator", "Name", "Budget (Million)", 
        "Jobs Created", "Completion Date", "Authority", "Status", 
        "General Status", "Last Updated"
    ]
    
    # Add headers with styling
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_num, project in enumerate(projects, 2):
        ws.cell(row=row_num, column=1, value=project.id)
        ws.cell(row=row_num, column=2, value=project.region.name)
        ws.cell(row=row_num, column=3, value=project.initiator)
        ws.cell(row=row_num, column=4, value=project.name)
        ws.cell(row=row_num, column=5, value=project.budget_million)
        ws.cell(row=row_num, column=6, value=project.jobs_created)
        ws.cell(row=row_num, column=7, value=project.completion_date.isoformat() if project.completion_date else None)
        ws.cell(row=row_num, column=8, value=project.authority.name)
        ws.cell(row=row_num, column=9, value=project.status.name)
        ws.cell(row=row_num, column=10, value=project.general_status)
        ws.cell(row=row_num, column=11, value=project.updated_at.isoformat() if project.updated_at else None)
    
    # Auto-adjust column widths
    for col_num, _ in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        # Set a minimum width
        ws.column_dimensions[column_letter].width = 15
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return Excel file with appropriate headers for direct download
    headers = {
        'Content-Disposition': 'attachment; filename="projects.xlsx"',
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Access-Control-Expose-Headers': 'Content-Disposition'
    }
    return Response(content=output.getvalue(), headers=headers)

# Update the read_project function to handle None current_user
@projects_router.get("/projects/{project_id}", response_model=schemas.ProjectDetail)
async def read_project(
    project_id: int,
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(conditional_auth)
):
    # This GET endpoint doesn't require authentication
    query = db.query(models.Project).filter(models.Project.id == project_id)
    
    if not include_deleted:
        query = filter_deleted(query, models.Project)
    
    project = query.first()
    
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Check if authenticated user has access to the project's region
    # For unauthenticated users, allow access to all projects
    if current_user is not None and not current_user.is_superadmin:
        if not check_region_access(current_user, project.region_id):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Not authorized to view this project"
            )
    
    return project

@projects_router.put("/projects/{project_id}", response_model=schemas.Project)
@projects_router.put("/projects/{project_id}/", response_model=schemas.Project)
async def update_project(
    project_id: int,
    project_update: schemas.ProjectUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    # Get the project
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.deleted_at == None
    ).first()
    
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Check if user has access to the project's region
    if not check_region_access(current_user, project.region_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to update this project"
        )
    
    # If updating region_id, check if user has access to the new region
    if project_update.region_id is not None and project_update.region_id != project.region_id:
        if not check_region_access(current_user, project_update.region_id):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Not authorized to move project to this region"
            )
        
        # Check if new region exists
        new_region = db.query(models.Region).filter(
            models.Region.id == project_update.region_id,
            models.Region.deleted_at == None
        ).first()
        if not new_region:
            raise HTTPException(status_code=404, detail="New region not found")
        
        project.region_id = project_update.region_id
    
    # Update other fields if provided
    if project_update.initiator is not None:
        project.initiator = project_update.initiator
    
    if project_update.name is not None:
        project.name = project_update.name
    
    if project_update.budget_million is not None:
        project.budget_million = project_update.budget_million
    
    if project_update.jobs_created is not None:
        project.jobs_created = project_update.jobs_created
    
    if project_update.completion_date is not None:
        project.completion_date = project_update.completion_date
    
    if project_update.authority_id is not None:
        # Check if authority exists
        authority = db.query(models.Authority).filter(
            models.Authority.id == project_update.authority_id,
            models.Authority.deleted_at == None
        ).first()
        if not authority:
            raise HTTPException(status_code=404, detail="Authority not found")
        
        project.authority_id = project_update.authority_id
    
    if project_update.status_id is not None:
        # Check if status exists
        status = db.query(models.Status).filter(
            models.Status.id == project_update.status_id,
            models.Status.deleted_at == None
        ).first()
        if not status:
            raise HTTPException(status_code=404, detail="Status not found")
        
        project.status_id = project_update.status_id
    
    if project_update.general_status is not None:
        project.general_status = project_update.general_status
    
    # updated_at will be automatically updated by SQLAlchemy
    
    db.commit()
    db.refresh(project)
    return project

# Add PATCH method for partial updates
@projects_router.patch("/projects/{project_id}", response_model=schemas.Project)
@projects_router.patch("/projects/{project_id}/", response_model=schemas.Project)
async def patch_project(
    project_id: int,
    project_update: schemas.ProjectUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    # Reuse the same implementation as PUT
    return await update_project(project_id, project_update, db, current_user)

@projects_router.delete("/projects/{project_id}", status_code=status.HTTP_204_NO_CONTENT)
@projects_router.delete("/projects/{project_id}/", status_code=status.HTTP_204_NO_CONTENT)
async def delete_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.deleted_at == None
    ).first()
    
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Check if user has access to the project's region
    if not check_region_access(current_user, project.region_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to delete this project"
        )
    
    # Soft delete
    project.soft_delete(db)
    return None

@projects_router.post("/projects/{project_id}/restore", response_model=schemas.Project)
async def restore_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.deleted_at != None
    ).first()
    
    if project is None:
        raise HTTPException(status_code=404, detail="Deleted project not found")
    
    # Check if user has access to the project's region
    if not check_region_access(current_user, project.region_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to restore this project"
        )
    
    # Restore project
    project.restore(db)
    return project

# Authority endpoints - Restored
@authorities_router.post("/authorities", response_model=schemas.Authority)
@authorities_router.post("/authorities/", response_model=schemas.Authority)
async def create_authority(
    authority: schemas.AuthorityCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    # Check if authority with the same name exists
    db_authority = db.query(models.Authority).filter(
        models.Authority.name == authority.name
    ).first()
    
    if db_authority and db_authority.deleted_at is None:
        raise HTTPException(status_code=400, detail="Authority with this name already exists")
    elif db_authority:
        # If authority was soft-deleted, restore it
        db_authority.deleted_at = None
    else:
        # Create new authority
        db_authority = models.Authority(**authority.dict())
        db.add(db_authority)
    
    db.commit()
    db.refresh(db_authority)
    return db_authority

# Update the read_authorities function to handle None current_user
@authorities_router.get("/authorities", response_model=List[schemas.Authority])
@authorities_router.get("/authorities/", response_model=List[schemas.Authority])
async def read_authorities(
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(conditional_auth)
):
    # This GET endpoint doesn't require authentication
    query = db.query(models.Authority)
    if not include_deleted:
        query = filter_deleted(query, models.Authority)
    authorities = query.all()
    return authorities

# Update the read_authority function to handle None current_user
@authorities_router.get("/authorities/{authority_id}", response_model=schemas.Authority)
async def read_authority(
    authority_id: int,
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(conditional_auth)
):
    # This GET endpoint doesn't require authentication
    query = db.query(models.Authority).filter(models.Authority.id == authority_id)
    if not include_deleted:
        query = filter_deleted(query, models.Authority)
    db_authority = query.first()
    
    if db_authority is None:
        raise HTTPException(status_code=404, detail="Authority not found")
    return db_authority

@authorities_router.put("/authorities/{authority_id}", response_model=schemas.Authority)
@authorities_router.put("/authorities/{authority_id}/", response_model=schemas.Authority)
async def update_authority(
    authority_id: int,
    authority_update: schemas.AuthorityCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    db_authority = db.query(models.Authority).filter(
        models.Authority.id == authority_id,
        models.Authority.deleted_at == None
    ).first()
    
    if db_authority is None:
        raise HTTPException(status_code=404, detail="Authority not found")
    
    # Check if name already exists
    existing_authority = db.query(models.Authority).filter(
        models.Authority.name == authority_update.name,
        models.Authority.id != authority_id,
        models.Authority.deleted_at == None
    ).first()
    if existing_authority:
        raise HTTPException(status_code=400, detail="Authority name already exists")
    
    db_authority.name = authority_update.name
    
    db.commit()
    db.refresh(db_authority)
    return db_authority

# Add PATCH method for partial updates
@authorities_router.patch("/authorities/{authority_id}", response_model=schemas.Authority)
@authorities_router.patch("/authorities/{authority_id}/", response_model=schemas.Authority)
async def patch_authority(
    authority_id: int,
    authority_update: schemas.AuthorityCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    # Reuse the same implementation as PUT
    return await update_authority(authority_id, authority_update, db, current_user)

@authorities_router.delete("/authorities/{authority_id}", status_code=status.HTTP_204_NO_CONTENT)
@authorities_router.delete("/authorities/{authority_id}/", status_code=status.HTTP_204_NO_CONTENT)
async def delete_authority(
    authority_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    db_authority = db.query(models.Authority).filter(
        models.Authority.id == authority_id,
        models.Authority.deleted_at == None
    ).first()
    
    if db_authority is None:
        raise HTTPException(status_code=404, detail="Authority not found")
    
    # Check if authority has associated projects
    projects = db.query(models.Project).filter(
        models.Project.authority_id == authority_id,
        models.Project.deleted_at == None
    ).count()
    
    if projects > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete authority with {projects} active projects"
        )
    
    # Soft delete
    db_authority.soft_delete(db)
    return None

@authorities_router.post("/authorities/{authority_id}/restore", response_model=schemas.Authority)
async def restore_authority(
    authority_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    db_authority = db.query(models.Authority).filter(
        models.Authority.id == authority_id,
        models.Authority.deleted_at != None
    ).first()
    
    if db_authority is None:
        raise HTTPException(status_code=404, detail="Deleted authority not found")
    
    # Restore authority
    db_authority.restore(db)
    return db_authority

# Status endpoints - Restored
@statuses_router.post("/statuses", response_model=schemas.Status)
@statuses_router.post("/statuses/", response_model=schemas.Status)
async def create_status(
    status: schemas.StatusCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    # Check if status with the same name exists
    db_status = db.query(models.Status).filter(
        models.Status.name == status.name
    ).first()
    
    if db_status and db_status.deleted_at is None:
        raise HTTPException(status_code=400, detail="Status with this name already exists")
    elif db_status:
        # If status was soft-deleted, restore it
        db_status.deleted_at = None
    else:
        # Create new status
        db_status = models.Status(**status.dict())
        db.add(db_status)
    
    db.commit()
    db.refresh(db_status)
    return db_status

# Update the read_statuses function to handle None current_user
@statuses_router.get("/statuses", response_model=List[schemas.Status])
@statuses_router.get("/statuses/", response_model=List[schemas.Status])
async def read_statuses(
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(conditional_auth)
):
    # This GET endpoint doesn't require authentication
    query = db.query(models.Status)
    if not include_deleted:
        query = filter_deleted(query, models.Status)
    statuses = query.all()
    return statuses

# Update the read_status function to handle None current_user
@statuses_router.get("/statuses/{status_id}", response_model=schemas.Status)
async def read_status(
    status_id: int,
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(conditional_auth)
):
    # This GET endpoint doesn't require authentication
    query = db.query(models.Status).filter(models.Status.id == status_id)
    if not include_deleted:
        query = filter_deleted(query, models.Status)
    db_status = query.first()
    
    if db_status is None:
        raise HTTPException(status_code=404, detail="Status not found")
    return db_status

@statuses_router.put("/statuses/{status_id}", response_model=schemas.Status)
@statuses_router.put("/statuses/{status_id}/", response_model=schemas.Status)
async def update_status(
    status_id: int,
    status_update: schemas.StatusCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    db_status = db.query(models.Status).filter(
        models.Status.id == status_id,
        models.Status.deleted_at == None
    ).first()
    
    if db_status is None:
        raise HTTPException(status_code=404, detail="Status not found")
    
    # Check if name already exists
    existing_status = db.query(models.Status).filter(
        models.Status.name == status_update.name,
        models.Status.id != status_id,
        models.Status.deleted_at == None
    ).first()
    if existing_status:
        raise HTTPException(status_code=400, detail="Status name already exists")
    
    db_status.name = status_update.name
    
    db.commit()
    db.refresh(db_status)
    return db_status

# Add PATCH method for partial updates
@statuses_router.patch("/statuses/{status_id}", response_model=schemas.Status)
@statuses_router.patch("/statuses/{status_id}/", response_model=schemas.Status)
async def patch_status(
    status_id: int,
    status_update: schemas.StatusCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    # Reuse the same implementation as PUT
    return await update_status(status_id, status_update, db, current_user)

@statuses_router.delete("/statuses/{status_id}", status_code=status.HTTP_204_NO_CONTENT)
@statuses_router.delete("/statuses/{status_id}/", status_code=status.HTTP_204_NO_CONTENT)
async def delete_status(
    status_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    db_status = db.query(models.Status).filter(
        models.Status.id == status_id,
        models.Status.deleted_at == None
    ).first()
    
    if db_status is None:
        raise HTTPException(status_code=404, detail="Status not found")
    
    # Check if status has associated projects
    projects = db.query(models.Project).filter(
        models.Project.status_id == status_id,
        models.Project.deleted_at == None
    ).count()
    
    if projects > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete status with {projects} active projects"
        )
    
    # Soft delete
    db_status.soft_delete(db)
    return None

@statuses_router.post("/statuses/{status_id}/restore", response_model=schemas.Status)
async def restore_status(
    status_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    db_status = db.query(models.Status).filter(
        models.Status.id == status_id,
        models.Status.deleted_at != None
    ).first()
    
    if db_status is None:
        raise HTTPException(status_code=404, detail="Deleted status not found")
    
    # Restore status
    db_status.restore(db)
    return db_status

# Root endpoint
@app.get("/", tags=["Root"])
async def root():
    return {"message": "Welcome to the Project Management API"}

# Initialize database endpoint
@app.post("/initialize-database", tags=["Administration"])
async def init_database(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_superadmin)
):
    """Initialize the database with data from the provided JSON file."""
    result = initialize_database(db)
    return {"message": "Database initialized successfully", "details": result}

# Include all routers
app.include_router(auth_router)
app.include_router(users_router)
app.include_router(regions_router)
app.include_router(projects_router)
app.include_router(authorities_router)
app.include_router(statuses_router)

# Customize OpenAPI schema to organize endpoints by tags
def custom_openapi():
    if app.openapi_schema:
        return app.openapi_schema
    
    openapi_schema = get_openapi(
        title=app.title,
        version=app.version,
        description=app.description,
        routes=app.routes,
    )
    
    # Customize the schema here if needed
    
    app.openapi_schema = openapi_schema
    return app.openapi_schema

app.openapi = custom_openapi

# Run the application
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
