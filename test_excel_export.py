"""
Script to test the Excel export functionality
"""
import os
import sys
import requests
import tempfile
import openpyxl
from dotenv import load_dotenv

# Add the parent directory to the path so we can import our modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

load_dotenv()

def test_excel_export():
    """Test the Excel export endpoint"""
    # Base URL for the API
    base_url = os.getenv("API_URL", "http://localhost:8000")
    
    # Test different scenarios
    test_cases = [
        {"name": "All projects (no parameters)", "url": f"{base_url}/projects/export"},
        {"name": "Projects by region", "url": f"{base_url}/projects/export?region_id=1"},
        {"name": "Projects by budget", "url": f"{base_url}/projects/export?budget_min=10&budget_max=100"},
        {"name": "Projects by status", "url": f"{base_url}/projects/export?status_id=1"},
        {"name": "Combined filters", "url": f"{base_url}/projects/export?region_id=1&status_id=1"}
    ]
    
    for test_case in test_cases:
        print(f"\nTesting: {test_case['name']}")
        print(f"URL: {test_case['url']}")
        
        try:
            # Make the request
            response = requests.get(test_case['url'])
            
            # Check status code
            print(f"Status code: {response.status_code}")
            if response.status_code != 200:
                print(f"Error: {response.text}")
                continue
            
            # Check headers
            print(f"Content-Type: {response.headers.get('Content-Type')}")
            print(f"Content-Disposition: {response.headers.get('Content-Disposition')}")
            
            # Save the file to a temporary location
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file.write(response.content)
                temp_path = temp_file.name
            
            # Try to open the Excel file
            try:
                wb = openpyxl.load_workbook(temp_path)
                ws = wb.active
                
                # Get some basic info about the file
                row_count = ws.max_row
                col_count = ws.max_column
                print(f"Excel file loaded successfully. Rows: {row_count}, Columns: {col_count}")
                
                # Print the first few rows as a sample
                if row_count > 1:
                    print("\nSample data:")
                    for row in range(1, min(4, row_count + 1)):
                        row_data = []
                        for col in range(1, col_count + 1):
                            cell_value = ws.cell(row=row, column=col).value
                            row_data.append(str(cell_value))
                        print(" | ".join(row_data))
                else:
                    print("No data rows found in the Excel file.")
                
                # Close the workbook
                wb.close()
                
                print("Excel file validation: SUCCESS")
            except Exception as e:
                print(f"Failed to open Excel file: {str(e)}")
            
            # Clean up the temporary file
            os.unlink(temp_path)
            
        except Exception as e:
            print(f"Request failed: {str(e)}")

if __name__ == "__main__":
    test_excel_export()
